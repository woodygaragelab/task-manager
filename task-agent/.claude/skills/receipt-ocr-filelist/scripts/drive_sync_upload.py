"""
Step6相当(receipt_filelist.xlsxと整理済ファイルのDriveへの反映)を、判断を伴わない
機械的な処理として1回のスクリプト実行にまとめたもの。旧来の「Claudeがcreate_file /
trash_fileをファイル数ぶん逐次呼ぶ」方式に代えて、Drive APIを直接(かつファイルの
アップロードは並列に)呼び出すことで高速化する。

使い方:
    python drive_sync_upload.py <client_code> [--work-dir DIR]
        [--receipt-folder-id ID] [--organized-folder-id ID]
        [--entries-json PATH] [--entries-payment-json PATH]
        [--state-file PATH] [--concurrency N]

前提(呼び出し前に用意されているもの。SKILL.md Step1/3/4/5相当):
    <work_dir>/receipt/receipt_filelist.xlsx   -- append_filelist.py で更新済み
    <work_dir>/renamed/<勘定科目名>/<ファイル名>  -- rename_and_save.py の出力(支払のみ)
    <work_dir>/entries.json                    -- 分類結果(支払/売上/給与/銀行通帳)
    <work_dir>/entries_payment.json            -- 支払分のみ(Step5でrename_and_save.pyに
        渡したものと同じ内容。無ければ<work_dir>/entries.jsonのうちclassificationが
        "支払"または未指定のものを使う)
    <work_dir>/new_files.json                  -- drive_sync_download.pyの出力
        (old_filelist_file_id / receipt_folder_id の取得元。無ければ
        --receipt-folder-id を明示指定すること。旧filelistが無い=新規作成の
        場合はold_filelist_file_idがnullのままでよい)

処理内容(**必ずこの順序で実行する**。receipt_filelist.xlsxは実ファイルのコピーが
確認できたものだけを反映するため、Driveへの反映は最後に行う):
    1. <work_dir>/renamed/配下の各ファイル(支払)を 整理済/支払/<勘定科目名>/ へ
       アップロード(同名ファイルがあれば新規アップロード後に旧ファイルをtrash)
    2. entries.jsonのうちclassificationが売上/給与/銀行通帳のファイルを、元の
       ファイル名のまま 整理済/<分類名>/ へアップロード(同上)
    3. receipt_filelist.xlsxを受領フォルダへアップロード → 成功後に旧ファイルをtrash
       (ローカルにreceipt_filelist.xlsxが無ければこのStepはスキップ)。1・2で
       アップロードに失敗したファイルがあれば、それらの行はアップロード前の
       receipt_filelist.xlsxから取り除いてから反映する(「ファイル名」列に記載が
       無い = 未処理」という次回実行時の判定と整合させ、失敗分を確実に再処理
       対象に戻すため)。
    個々のファイルのアップロードに失敗しても処理全体は止めず、失敗分をスキップして
    最後にまとめて報告する。

出力(stdout, JSON): {"payment": {...}, "others": {...}, "filelist": {...},
"failures": [...]}

タイミングログ(stderr): [timing] phase=upload_filelist / trash_old_filelist /
search_or_create_folder(extra: name) / upload_file(extra: filename) /
trash_old_file(extra: filename) / total
"""
import argparse
import json
import os
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed

from drive_sync_common import (
    Timing,
    execute_with_retry,
    get_drive_service,
    log,
    resolve_client_folders,
    search_or_create_folder,
    upload_file_replacing,
)
from rename_and_save import safe_component

FILELIST_NAME = "receipt_filelist.xlsx"
COPY_AS_IS_CLASSIFICATIONS = {"売上", "給与", "銀行通帳"}


def get_organized_root_id(timing: Timing, receipt_folder_id: str, organized_folder_id: str) -> str:
    if organized_folder_id:
        return organized_folder_id
    service = get_drive_service()
    meta = execute_with_retry(service.files().get(fileId=receipt_folder_id, fields="parents"))
    parents = meta.get("parents") or []
    if not parents:
        raise RuntimeError(f"receipt_folder_id={receipt_folder_id!r} に親フォルダがありません")
    client_folder_id = parents[0]
    return search_or_create_folder(client_folder_id, "整理済", timing)


def load_entries(path):
    if not os.path.exists(path):
        return []
    return json.load(open(path, encoding="utf-8"))


def payment_dest_name(entry: dict) -> tuple:
    """rename_and_save.pyと同じ命名規則で(勘定科目名フォルダ名, リネーム後ファイル名)を
    再計算する(rename_and_save.py自体は無改造のまま使うため、命名関数だけをimportして
    再利用する)。"""
    account = safe_component(entry.get("account_name"), "未分類")
    vendor = safe_component(entry.get("vendor"), "取引先不明")
    date_str = (entry.get("date") or "日付不明").replace("-", "")
    amount = entry.get("amount")
    amount_str = f"{amount}円" if amount is not None else "金額不明"
    ext = os.path.splitext(entry["filename"])[1].lower()
    return account, f"{account}_{date_str}_{amount_str}_{vendor}{ext}"


def upload_payment_files(timing: Timing, renamed_dir: str, organized_root_id: str, concurrency: int, payment_entries: list):
    tasks = []
    for e in payment_entries:
        account, dest_name = payment_dest_name(e)
        local_path = os.path.join(renamed_dir, account, dest_name)
        if not os.path.exists(local_path):
            log(f"[skip] renamedファイルが見つかりません: {local_path}")
            continue
        tasks.append((account, dest_name, local_path, e["filename"]))

    results, failures = [], []

    def _run(account, filename, local_path, original_filename):
        payment_root = search_or_create_folder(organized_root_id, "支払", timing)
        account_folder = search_or_create_folder(payment_root, account, timing)
        with timing.phase("upload_file", {"filename": filename}):
            r = upload_file_replacing(local_path, account_folder, filename)
        return {"account": account, "filename": filename, "original_filename": original_filename, **r}

    with ThreadPoolExecutor(max_workers=concurrency) as pool:
        future_to_task = {pool.submit(_run, *t): t for t in tasks}
        for future in as_completed(future_to_task):
            account, filename, _, original_filename = future_to_task[future]
            try:
                results.append(future.result())
            except Exception as e:
                log(f"[error] upload failed account={account!r} filename={filename!r}: {e}")
                failures.append(
                    {
                        "filename": filename,
                        "account": account,
                        "original_filename": original_filename,
                        "error": str(e),
                    }
                )

    return {"count": len(results), "results": results, "failures": failures}


def upload_other_files(
    timing: Timing, receipt_dir: str, organized_root_id: str, concurrency: int, other_entries: list
):
    tasks = []
    for e in other_entries:
        classification = e["classification"]
        filename = e["filename"]
        src = os.path.join(receipt_dir, *filename.replace("\\", "/").split("/"))
        if not os.path.exists(src):
            log(f"[skip] 元ファイルが見つかりません: {src}")
            continue
        dest_name = os.path.basename(filename.replace("\\", "/"))
        tasks.append((classification, dest_name, src, filename))

    results, failures = [], []

    def _run(classification, dest_name, src, original_filename):
        dest_folder = search_or_create_folder(organized_root_id, classification, timing)
        with timing.phase("upload_file", {"filename": dest_name}):
            r = upload_file_replacing(src, dest_folder, dest_name)
        return {"classification": classification, "filename": dest_name, "original_filename": original_filename, **r}

    with ThreadPoolExecutor(max_workers=concurrency) as pool:
        future_to_task = {pool.submit(_run, *t): t for t in tasks}
        for future in as_completed(future_to_task):
            classification, dest_name, _, original_filename = future_to_task[future]
            try:
                results.append(future.result())
            except Exception as e:
                log(f"[error] upload failed classification={classification!r} filename={dest_name!r}: {e}")
                failures.append(
                    {
                        "filename": dest_name,
                        "classification": classification,
                        "original_filename": original_filename,
                        "error": str(e),
                    }
                )

    return {"count": len(results), "results": results, "failures": failures}


def remove_failed_rows(local_path: str, failed_filenames: set):
    """アップロードに失敗したファイルの行をreceipt_filelist.xlsxから取り除く
    (「ファイル名」列に記載が無い=未処理、という次回実行時の判定に合わせるため。
    アップロード前のローカルファイルを直接書き換える)。"""
    if not failed_filenames or not os.path.exists(local_path):
        return
    from openpyxl import load_workbook

    normalized_failed = {f.replace("\\", "/") for f in failed_filenames}
    wb = load_workbook(local_path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    try:
        filename_col = header.index("ファイル名") + 1
    except ValueError:
        return
    rows_to_delete = []
    for row in ws.iter_rows(min_row=2):
        cell = row[filename_col - 1]
        if cell.value and str(cell.value).replace("\\", "/") in normalized_failed:
            rows_to_delete.append(cell.row)
    for row_idx in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_idx)
    if rows_to_delete:
        wb.save(local_path)
        log(f"[remove_failed_rows] {len(rows_to_delete)}行を削除(次回再処理対象に戻す)")


def upload_filelist(timing: Timing, receipt_dir: str, receipt_folder_id: str, old_filelist_file_id, failed_filenames: set):
    local_path = os.path.join(receipt_dir, FILELIST_NAME)
    if not os.path.exists(local_path):
        log("[upload_filelist] ローカルにreceipt_filelist.xlsxが無いためスキップ")
        return {"uploaded": False}
    remove_failed_rows(local_path, failed_filenames)
    with timing.phase("upload_filelist"):
        result = upload_file_replacing(local_path, receipt_folder_id, FILELIST_NAME)
    if result["replaced"] is False and old_filelist_file_id:
        # upload_file_replacingは名前一致で旧ファイルを検索するため通常は
        # 見つかるはずだが、念のためold_filelist_file_idが分かっていれば
        # 明示的にもtrashしておく(名前が変わっていた場合の取りこぼし対策)。
        with timing.phase("trash_old_filelist"):
            service = get_drive_service()
            execute_with_retry(
                service.files().update(fileId=old_filelist_file_id, body={"trashed": True})
            )
        result["replaced"] = True
    log(f"[upload_filelist] file_id={result['file_id']} replaced={result['replaced']}")
    return {"uploaded": True, **result}


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("client_code")
    parser.add_argument("--work-dir", default=None, help="デフォルト: /tmp/{client_code}")
    parser.add_argument("--receipt-folder-id", default=None)
    parser.add_argument("--organized-folder-id", default=None, help="「整理済」フォルダIDを直接指定(省略時は受領フォルダの親から解決)")
    parser.add_argument("--entries-json", default=None, help="デフォルト: <work_dir>/entries.json")
    parser.add_argument("--entries-payment-json", default=None, help="デフォルト: <work_dir>/entries_payment.json(無ければ--entries-jsonのうちclassificationが支払/未指定のものを使う)")
    parser.add_argument("--state-file", default=None, help="デフォルト: <work_dir>/new_files.json")
    parser.add_argument("--concurrency", type=int, default=5)
    args = parser.parse_args()

    work_dir = args.work_dir or os.path.join("/tmp", args.client_code)
    receipt_dir = os.path.join(work_dir, "receipt")
    renamed_dir = os.path.join(work_dir, "renamed")
    entries_json_path = args.entries_json or os.path.join(work_dir, "entries.json")
    entries_payment_json_path = args.entries_payment_json or os.path.join(work_dir, "entries_payment.json")
    state_file = args.state_file or os.path.join(work_dir, "new_files.json")

    receipt_folder_id = args.receipt_folder_id
    old_filelist_file_id = None
    if os.path.exists(state_file):
        state = json.load(open(state_file, encoding="utf-8"))
        receipt_folder_id = receipt_folder_id or state.get("receipt_folder_id")
        old_filelist_file_id = state.get("old_filelist_file_id")

    if not receipt_folder_id:
        client = resolve_client_folders(args.client_code)
        receipt_folder_id = client.get("receiptFolderId")
        if not receipt_folder_id:
            print(
                json.dumps(
                    {"error": f"client_code={args.client_code!r} にreceiptFolderIdが未設定です"},
                    ensure_ascii=False,
                )
            )
            sys.exit(1)

    all_entries = load_entries(entries_json_path)
    other_entries = [e for e in all_entries if e.get("classification") in COPY_AS_IS_CLASSIFICATIONS]
    if os.path.exists(entries_payment_json_path):
        payment_entries = load_entries(entries_payment_json_path)
    else:
        payment_entries = [e for e in all_entries if e.get("classification", "支払") == "支払"]

    timing = Timing()
    organized_root_id = get_organized_root_id(timing, receipt_folder_id, args.organized_folder_id)

    # receipt_filelist.xlsxのDriveへの反映は、実ファイルのコピーが終わってから
    # 最後に行う(失敗した分の行を取り除いてからアップロードするため)。
    payment_result = upload_payment_files(timing, renamed_dir, organized_root_id, args.concurrency, payment_entries)
    other_result = upload_other_files(timing, receipt_dir, organized_root_id, args.concurrency, other_entries)

    failures = payment_result.get("failures", []) + other_result.get("failures", [])
    failed_filenames = {f["original_filename"] for f in failures}
    filelist_result = upload_filelist(timing, receipt_dir, receipt_folder_id, old_filelist_file_id, failed_filenames)

    total_ms = timing.total()

    result = {
        "client_code": args.client_code,
        "filelist": filelist_result,
        "payment": {"count": payment_result["count"], "results": payment_result["results"]},
        "others": {"count": other_result["count"], "results": other_result["results"]},
        "failures": failures,
        "elapsed_ms": total_ms,
    }
    print(json.dumps(result, ensure_ascii=False, indent=2))
    if failures:
        sys.exit(1)


if __name__ == "__main__":
    main()
