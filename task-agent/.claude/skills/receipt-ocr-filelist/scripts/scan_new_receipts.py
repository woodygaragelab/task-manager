"""
receipt フォルダ(サブフォルダ含む)内の画像ファイルを調べ、receipt_filelist.xlsx に
まだ載っていない新規ファイルだけを JSON で出力する。

使い方:
    python scan_new_receipts.py <receiptフォルダ> <receipt_filelist.xlsxのパス>

出力 (stdout, JSON):
    [
      {"filename": "IMG_20260611_091550.jpg", "taken_at": "2026-06-11T09:15:50",
       "size_mb": 3.34, "type": "JPEG画像"},
      {"filename": "2026年6月\\IMG_20260605_101200.jpg", "taken_at": "2026-06-05T10:12:00",
       "size_mb": 2.10, "type": "JPEG画像"},
      ...
    ]
"filename" は receipt フォルダを起点とした相対パス(サブフォルダ内のファイルは
"サブフォルダ名\ファイル名" の形式)。receipt_filelist.xlsx の「ファイル名」列には
この値をそのまま書き込むので、実ファイルを開く際は
os.path.join(<receiptフォルダ>, filename) で絶対パスを組み立てること。
taken_at は IMG_YYYYMMDD_HHMMSS 形式のファイル名(サブフォルダ部分は除く)からのみ
抽出する。パターンに合わない場合は null になるので、その場合は画像内の日付などから判断する。
"""
import sys, os, re, json, datetime
from openpyxl import load_workbook

EXT_TYPE = {
    ".jpg": "JPEG画像", ".jpeg": "JPEG画像", ".png": "PNG画像",
    ".heic": "HEIC画像", ".pdf": "PDFファイル",
}

def main():
    receipt_dir, filelist_path = sys.argv[1], sys.argv[2]

    existing = set()
    if os.path.exists(filelist_path):
        wb = load_workbook(filelist_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1]:
                existing.add(row[1])

    new_files = []
    for dirpath, dirnames, filenames in os.walk(receipt_dir):
        dirnames.sort()
        for f in sorted(filenames):
            ext = os.path.splitext(f)[1].lower()
            if ext not in EXT_TYPE:
                continue
            full_path = os.path.join(dirpath, f)
            rel_path = os.path.relpath(full_path, receipt_dir)
            if rel_path in existing:
                continue
            size_mb = round(os.path.getsize(full_path) / 1024 / 1024, 2)
            m = re.match(r"IMG_(\d{8})_(\d{6})", f)
            taken_at = None
            if m:
                taken_at = datetime.datetime.strptime(
                    m.group(1) + m.group(2), "%Y%m%d%H%M%S"
                ).isoformat()
            new_files.append({
                "filename": rel_path,
                "taken_at": taken_at,
                "size_mb": size_mb,
                "type": EXT_TYPE[ext],
            })

    print(json.dumps(new_files, ensure_ascii=False, indent=2))

if __name__ == "__main__":
    main()
