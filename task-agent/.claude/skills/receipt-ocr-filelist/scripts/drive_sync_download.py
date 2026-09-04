"""
Step1相当(Drive上の新規ファイルの検出とダウンロード)を、判断を伴わない機械的な
処理として1回のスクリプト実行にまとめたもの。旧来の「Claudeがsearch_files /
download_file_contentをファイル数ぶん逐次呼ぶ」方式に代えて、Drive APIを直接
(かつ新規ファイルのダウンロードは並列に)呼び出すことで高速化する。

使い方:
    python drive_sync_download.py <client_code> [--work-dir DIR]
        [--receipt-folder-id ID] [--concurrency N]

client_code はtaskmanagerのクライアントマスタ(TaskClients)のclientCode。
--receipt-folder-id を指定した場合はクライアントマスタの参照を省略できる
(DynamoDBの環境変数が無いローカル実行・テスト向け)。

出力(stdout, JSON): scan_new_receipts.pyと同じ"filename/taken_at/size_mb/type"を
持つ新規ファイルのリストに、file_id(Drive上のID)とreceipt_folder_id/
old_filelist_file_id(Step5相当のアップロード側で使う)を加えたもの。
同じ内容を <work_dir>/new_files.json にも書き出す。

タイミングログ(stderr): [timing] phase=list_files / download_filelist / diff /
download_file(extra: filename) / total
"""
import argparse
import json
import os
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed

from googleapiclient.http import MediaIoBaseDownload

from drive_sync_common import (
    FOLDER_MIME_TYPE,
    Timing,
    execute_with_retry,
    get_drive_service,
    log,
    resolve_client_folders,
)
from scan_new_receipts import EXT_TYPE

FILELIST_NAME = "receipt_filelist.xlsx"
TAKEN_AT_RE = re.compile(r"IMG_(\d{8})_(\d{6})")


def list_receipt_tree(timing: Timing, root_folder_id: str):
    """受領フォルダ配下を再帰的に列挙する。戻り値は
    (ファイル一覧[{id,name,mimeType,size,rel_path}], receipt_filelist.xlsxの
    メタ情報 or None)。"""
    files = []
    filelist_entry = None
    with timing.phase("list_files"):
        service = get_drive_service()
        queue = [(root_folder_id, "")]
        while queue:
            parent_id, prefix = queue.pop(0)
            page_token = None
            while True:
                resp = execute_with_retry(
                    service.files().list(
                        q=f"'{parent_id}' in parents and trashed = false",
                        fields="nextPageToken, files(id, name, mimeType, size)",
                        pageSize=200,
                        pageToken=page_token,
                    )
                )
                for f in resp.get("files", []):
                    if f["mimeType"] == FOLDER_MIME_TYPE:
                        queue.append((f["id"], f"{prefix}{f['name']}/"))
                        continue
                    if prefix == "" and f["name"] == FILELIST_NAME:
                        filelist_entry = f
                        continue
                    files.append({**f, "rel_path": f"{prefix}{f['name']}"})
                page_token = resp.get("nextPageToken")
                if not page_token:
                    break
    log(f"[list_files] {len(files)} file(s), filelist={'found' if filelist_entry else 'none'}")
    return files, filelist_entry


def download_existing_filelist(timing: Timing, filelist_entry: dict, dest_path: str):
    with timing.phase("download_filelist"):
        service = get_drive_service()
        request = service.files().get_media(fileId=filelist_entry["id"])
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        with open(dest_path, "wb") as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                _, done = execute_with_retry_downloader(downloader)
    log(f"[download_filelist] {filelist_entry['name']} -> {dest_path}")


def execute_with_retry_downloader(downloader, max_retries: int = 5, base_delay: float = 1.0):
    """MediaIoBaseDownload.next_chunk()はHttpErrorを投げうるが引数を取らないため
    execute_with_retryのシグネチャに合わないので専用の薄いリトライを用意する。"""
    import random
    import time as _time

    from googleapiclient.errors import HttpError

    for attempt in range(max_retries + 1):
        try:
            return downloader.next_chunk()
        except HttpError as e:
            status = e.resp.status if e.resp is not None else None
            if status not in (429, 500, 502, 503, 504) or attempt == max_retries:
                raise
            delay = base_delay * (2**attempt) + random.uniform(0, 0.5)
            log(f"[retry] download status={status} attempt={attempt + 1}/{max_retries} wait={delay:.1f}s")
            _time.sleep(delay)


def diff_new_files(timing: Timing, remote_files: list, local_filelist_path: str) -> list:
    """既存receipt_filelist.xlsxの「ファイル名」列と突き合わせ、未記載のものだけ
    新規ファイルとして返す(scan_new_receipts.pyと同じロジック)。"""
    with timing.phase("diff"):
        existing = set()
        if os.path.exists(local_filelist_path):
            from openpyxl import load_workbook

            wb = load_workbook(local_filelist_path, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[1]:
                    existing.add(str(row[1]).replace("\\", "/"))

        new_files = []
        for f in remote_files:
            ext = os.path.splitext(f["name"])[1].lower()
            if ext not in EXT_TYPE:
                continue
            if f["rel_path"] in existing:
                continue
            size_mb = round(int(f.get("size", 0)) / 1024 / 1024, 2)
            m = TAKEN_AT_RE.match(f["name"])
            taken_at = None
            if m:
                import datetime

                taken_at = datetime.datetime.strptime(
                    m.group(1) + m.group(2), "%Y%m%d%H%M%S"
                ).isoformat()
            new_files.append(
                {
                    "filename": f["rel_path"],
                    "taken_at": taken_at,
                    "size_mb": size_mb,
                    "type": EXT_TYPE[ext],
                    "file_id": f["id"],
                }
            )
    log(f"[diff] {len(new_files)} new file(s) out of {len(remote_files)}")
    return new_files


def download_one(timing: Timing, entry: dict, receipt_dir: str):
    dest = os.path.join(receipt_dir, *entry["filename"].split("/"))
    with timing.phase("download_file", {"filename": entry["filename"]}):
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        service = get_drive_service()
        request = service.files().get_media(fileId=entry["file_id"])
        with open(dest, "wb") as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                _, done = execute_with_retry_downloader(downloader)
    return dest


def download_new_files(timing: Timing, new_files: list, receipt_dir: str, concurrency: int):
    if not new_files:
        return
    errors = []
    with ThreadPoolExecutor(max_workers=concurrency) as pool:
        future_to_entry = {
            pool.submit(download_one, timing, entry, receipt_dir): entry for entry in new_files
        }
        for future in as_completed(future_to_entry):
            entry = future_to_entry[future]
            try:
                future.result()
            except Exception as e:
                log(f"[error] download failed filename={entry['filename']!r}: {e}")
                errors.append({"filename": entry["filename"], "error": str(e)})
    if errors:
        raise RuntimeError(
            f"{len(errors)}件のダウンロードに失敗しました: "
            + json.dumps(errors, ensure_ascii=False)
        )


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("client_code")
    parser.add_argument("--work-dir", default=None, help="デフォルト: /tmp/{client_code}")
    parser.add_argument("--receipt-folder-id", default=None)
    parser.add_argument("--concurrency", type=int, default=5)
    args = parser.parse_args()

    work_dir = args.work_dir or os.path.join("/tmp", args.client_code)
    receipt_dir = os.path.join(work_dir, "receipt")
    os.makedirs(receipt_dir, exist_ok=True)

    receipt_folder_id = args.receipt_folder_id
    if not receipt_folder_id:
        client = resolve_client_folders(args.client_code)
        receipt_folder_id = client.get("receiptFolderId")
        if not receipt_folder_id:
            print(
                json.dumps(
                    {"error": f"client_code={args.client_code!r} にreceiptFolderIdが未設定です"},
                    ensure_ascii=False,
                )
            )
            sys.exit(1)

    timing = Timing()
    remote_files, filelist_entry = list_receipt_tree(timing, receipt_folder_id)

    local_filelist_path = os.path.join(receipt_dir, FILELIST_NAME)
    if filelist_entry:
        download_existing_filelist(timing, filelist_entry, local_filelist_path)

    new_files = diff_new_files(timing, remote_files, local_filelist_path)
    download_new_files(timing, new_files, receipt_dir, args.concurrency)

    total_ms = timing.total()

    result = {
        "client_code": args.client_code,
        "receipt_folder_id": receipt_folder_id,
        "old_filelist_file_id": filelist_entry["id"] if filelist_entry else None,
        "new_files": new_files,
        "elapsed_ms": total_ms,
    }
    state_path = os.path.join(work_dir, "new_files.json")
    with open(state_path, "w", encoding="utf-8") as fh:
        json.dump(result, fh, ensure_ascii=False, indent=2)

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
