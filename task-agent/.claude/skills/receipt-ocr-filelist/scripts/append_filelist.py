"""
領収書から読み取った内容を receipt_filelist.xlsx に追記する。
ファイルが無ければヘッダー付きで新規作成し、旧フォーマット(区分/支払方法/消費税額/仕訳No/
勘定科目名列が無い)の場合はヘッダー行に不足列を追加する(既存データ行はそのまま)。

使い方:
    python append_filelist.py <entries.jsonのパス> <receipt_filelist.xlsxのパス>

entries.json は次の形式のリスト(1要素=1領収書):
[
  {
    "filename": "IMG_20260611_091550.jpg",
    "taken_at": "2026-06-11T09:15:50",   // null可。null なら撮影日時セルは空欄
    "size_mb": 3.34,
    "type": "JPEG画像",
    "date": "2026-02-04",                // 領収書に記載の日付 (YYYY-MM-DD)
    "amount": 5800,
    "vendor": "東京空港交通(株)・京成バス(株)",
    "content": "リムジンバス 成田空港T2→新浦安駅(片道乗車券・大人2名)",
    "category": "領収書",                // "領収書" or "請求書"
    "payment_method": "現金",            // 現金/カード/振込/不明
    "tax_amount": 527,                   // 消費税額の記載(無ければ null)
    "account_name": "旅費交通費"          // master/TKCマスタ.xlsx の勘定科目コード表と照合した
                                          // 参考の勘定科目名(借方相当)。判断できなければ null。
                                          // ※ここでは仕訳(借方/貸方コード)は確定させない。
                                          //   最終的な科目コードは convert-pattern-b /
                                          //   receipt-to-journal スキル側で決定する。
  }
]
新規行の「仕訳No」列は常に空欄のまま(まだ仕訳化されていない印。receipt-to-journal スキルが
処理後に値を入れる)。
"""
import sys, os, json, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment

HEADERS = ["No.", "ファイル名", "撮影日時", "サイズ(MB)", "種類", "日付", "金額", "取引先", "内容",
           "区分", "支払方法", "消費税額", "仕訳No", "勘定科目名"]
# 既存ファイルのヘッダーを補完する際に、このスキルが追加してよい列(「仕訳No」は含めない)。
# 「仕訳No」列は receipt-to-journal スキル側が、移行時の過去分バックフィルとセットで
# 追加する責任を持つ。ここで先に追加してしまうと、その移行処理が二度と行われなくなり、
# 過去の行がいつまでも「未仕訳」に見えてしまう。
HEADERS_TO_BACKFILL = [h for h in HEADERS if h != "仕訳No"]


def open_or_create(path):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb["receipt"] if "receipt" in wb.sheetnames else wb.active
        existing_headers = [c.value for c in ws[1] if c.value is not None]
        for h in HEADERS_TO_BACKFILL:
            if h not in existing_headers:
                col_idx = ws.max_column + 1
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = Font(name="Arial", bold=True)
                cell.alignment = Alignment(horizontal="center")
                existing_headers.append(h)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "receipt"
        ws.append(HEADERS)
        for c in ws[1]:
            c.font = Font(name="Arial", bold=True)
            c.alignment = Alignment(horizontal="center")
    return wb, ws


def next_seq(ws, col_idx):
    max_no = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[col_idx - 1] if len(row) >= col_idx else None
        if v and isinstance(v, (int, float)):
            max_no = max(max_no, int(v))
    return max_no + 1


def main():
    entries_path, filelist_path = sys.argv[1], sys.argv[2]
    entries = json.load(open(entries_path, encoding="utf-8"))

    wb, ws = open_or_create(filelist_path)
    header_row = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(header_row) if h}
    no = next_seq(ws, col["No."])

    for e in entries:
        taken_at = datetime.datetime.fromisoformat(e["taken_at"]) if e.get("taken_at") else None
        r = ws.max_row + 1
        ws.cell(row=r, column=col["No."], value=no)
        ws.cell(row=r, column=col["ファイル名"], value=e["filename"])
        if taken_at:
            c = ws.cell(row=r, column=col["撮影日時"], value=taken_at)
            c.number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row=r, column=col["サイズ(MB)"], value=e["size_mb"]).number_format = "0.00"
        ws.cell(row=r, column=col["種類"], value=e["type"])
        if e.get("date"):
            c = ws.cell(row=r, column=col["日付"], value=datetime.datetime.strptime(e["date"], "%Y-%m-%d"))
            c.number_format = "yyyy-mm-dd"
        if e.get("amount") is not None:
            ws.cell(row=r, column=col["金額"], value=e["amount"]).number_format = "#,##0"
        ws.cell(row=r, column=col["取引先"], value=e.get("vendor"))
        ws.cell(row=r, column=col["内容"], value=e.get("content"))
        ws.cell(row=r, column=col["区分"], value=e.get("category"))
        ws.cell(row=r, column=col["支払方法"], value=e.get("payment_method"))
        if e.get("tax_amount") is not None:
            ws.cell(row=r, column=col["消費税額"], value=e["tax_amount"]).number_format = "#,##0"
        if "勘定科目名" in col:
            ws.cell(row=r, column=col["勘定科目名"], value=e.get("account_name"))
        # 「仕訳No」は空欄のまま(receipt-to-journal スキルが後で埋める)
        for cell in ws[r]:
            cell.font = Font(name="Arial")
        no += 1

    wb.save(filelist_path)
    print(f"receipt_filelist: +{len(entries)}行 -> {filelist_path}")


if __name__ == "__main__":
    main()
